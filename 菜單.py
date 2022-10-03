import openpyxl
import tkinter as tk
import tkinter.font as tkFont
from tkinter import ttk
from openpyxl.styles import PatternFill, Font

date = input('請輸入期(**/**):')

class Action(tk.Frame):
    x = 3


    def __init__(self):
        tk.Frame.__init__(self)
        self.grid()
        self.master.geometry('600x600')  # 設定視窗初始大小
        self.master.minsize(600, 600)  # 最大
        self.master.maxsize(600, 600)  # 最小
        self.master.title('菜單吾皇萬福金安')  # 名稱
        self.createWidgets()  # 其他東東
        self.ori = ''


    def createWidgets(self):
        # 字型設定
        f1 = tkFont.Font(size=30, family="Courier New")
        f2 = tkFont.Font(size=25, family="Courier New")
        f3 = tkFont.Font(size=20, family="Courier New")
        f4 = tkFont.Font(size=16, family="Courier New")
        # 白色背景
        self.background = tk.Canvas(self, height=618, width=1000, bg='white').pack()
        # 按鈕
        tk.Button(self, text="開始", height=1, width=4, bg='#ffcc00', font=f3, command=self.click).place(x=250, y=350)
        tk.Button(self, text="加總", height=1, width=4, bg='#009962', font=f3, command=self.add).place(x=250, y=450)
        tk.Button(self, text="編號", height=1, width=4, bg='#FF2D2D', font=f3, command=self.num).place(x=250, y=530)
        # 標題
        self.name_label = tk.Label(self, text="人名:", height=1, width=10, bg='white', font=f2).place(x=17, y=50)
        self.time_label = tk.Label(self, text="時間:", height=1, width=10, bg='white', font=f2).place(x=17, y=100)
        self.local_label = tk.Label(self, text="地點:", height=1, width=10, bg='white', font=f2).place(x=17, y=150)
        self.veg_label = tk.Label(self, text="菜:", height=1, width=10, bg='white', font=f2).place(x=17, y=200)
        self.quan_label = tk.Label(self, text="量:", height=1, width=10, bg='white', font=f2).place(x=17, y=250)
        self.hit_label = tk.Label(self, text="重量:", height=1, width=10, bg='white', font=f2).place(x=17, y=300)
        # 輸入
        self.name = tk.Entry(self, textvariable=tk.StringVar(), bg='pink', font=f3)  # 老師的
        self.name.place(x=200, y=50)  # 位置
        self.time = tk.Entry(self, textvariable=tk.StringVar(), bg='pink', font=f3)
        self.time.place(x=200, y=100)
        self.local = tk.Entry(self, textvariable=tk.StringVar(), bg='pink', font=f3)
        self.local.place(x=200, y=150)
        self.veg = tk.Entry(self, textvariable=tk.StringVar(), bg='pink', font=f3)
        self.veg.place(x=200, y=200)
        self.quan = tk.Entry(self, textvariable=tk.StringVar(), bg='pink', font=f3)
        self.quan.place(x=200, y=250)
        self.hit = tk.Entry(self, textvariable=tk.StringVar(), bg='pink', font=f3)
        self.hit.place(x=200, y=300)


    def num(self):
        wb0 = openpyxl.load_workbook('D:/菜單/0227菜單.xlsx')
        font0 = Font(size=20, bold=True)
        for sheet in wb0:
            x = 3
            if sheet.title != '總單':
                while True:
                    if sheet.cell(x, 2).value is None:
                        break
                    sheet.cell(row = x, column = 1, value = (str(x-2) + '.   '))
                    sheet.cell(x, 1).font = font0
                    x += 1
        wb0.save('D:/菜單/0227菜單.xlsx')


    def click(self):  # 點擊程序
        time = self.time.get()
        name = self.name.get()
        veg = self.veg.get()
        local = self.local.get()
        hit = self.hit.get()
        quan = self.quan.get()
        hit = self.hit.get()
        self.wb0 = openpyxl.load_workbook('D:/菜單/0227菜單.xlsx')
        font0 = Font(size=20, bold=True)
        if name == self.ori:
            sheet = self.wb0[time + self.ori]
            Action.x += 1
            sheet.cell(row = Action.x, column = 2, value = veg)
            sheet.cell(row = Action.x, column = 3, value = quan)
        else:
            Action.x = 3
            self.ori = name
            self.wb0.create_sheet(title=(time + self.ori))  # 新增工作表
            self.wb0.save('D:/菜單/0227菜單.xlsx')
            self.wb0 = openpyxl.load_workbook('D:/菜單/0227菜單.xlsx')
            sheet = self.wb0[time + self.ori]  # 取得第一個工作表
            sheet.cell(row = 1, column = 2, value = name)
            sheet.cell(row = 1, column = 3, value = time)
            sheet.cell(row = 1, column = 4, value = local)
            sheet.cell(row = 2, column = 2, value = '菜名')
            sheet.cell(row = 2, column = 3, value = '數量')
            sheet.cell(row = 2, column = 4, value = '     斤,兩')
            sheet.cell(row = 2, column = 5, value = '單價')
            sheet.cell(row = 2, column = 6, value = '金額')
            pattern_fill = PatternFill(fill_type="solid", fgColor="FFFF37")  # 这个 color 是 RGB 的 16 进制表示
            for i in range(2, 7):
               sheet.cell(row = 2, column = i).fill = pattern_fill
            sheet.cell(row = Action.x, column = 2, value = veg)
            sheet.cell(row = Action.x, column = 3, value = quan)
            sheet.cell(row = 1, column = 1, value = date)
            sheet.cell(1, 1).font = font0
            sheet.cell(1, 2).font = font0
            sheet.cell(1, 3).font = font0
            sheet.cell(1, 4).font = font0
            sheet.cell(2, 2).font = font0
            sheet.cell(2, 3).font = font0
            sheet.cell(2, 4).font = font0
            sheet.cell(2, 5).font = font0
            sheet.cell(2, 6).font = font0
        if hit != '':
            sheet.cell(row = Action.x, column = 4, value = hit)
        sheet.cell(Action.x, 2).font = font0
        sheet.cell(Action.x, 3).font = font0
        sheet.cell(Action.x, 4).font = font0
        self.wb0.save('D:/菜單/0227菜單.xlsx')


    def add(self):
        font0 = Font(size=20, bold=True)
        hit_dict = {}
        veg_dict = {}
        self.wb0 = openpyxl.load_workbook('D:/菜單/0227菜單.xlsx')
        self.wb0.create_sheet('總單')
        self.wb0.save('D:/菜單/0227菜單.xlsx')
        self.wb0 = openpyxl.load_workbook('D:/菜單/0227菜單.xlsx')

        for sheet in self.wb0:
            x = 3
            if sheet.title == '總單':
                sheet.cell(row = 2, column = 2, value = '總單')
                sheet.cell(row = 2, column = 3, value = '數量')
                sheet.cell(row = 2, column = 4, value = '備註')
                sheet.cell(2, 2).font = font0
                sheet.cell(2, 3).font = font0
                sheet.cell(2, 4).font = font0
                pattern_fill = PatternFill(fill_type="solid", fgColor="FFFF37")  # 这个 color 是 RGB 的 16 进制表示
                for i in range(2, 5):
                    sheet.cell(row = 2, column = i).fill = pattern_fill
                key = list(veg_dict.keys())
                for i in range(len(key)):
                    gg = 0
                    sheet.cell(row = (x + i), column = 2, value = key[i])
                    sheet.cell(row = (x + i), column = 3, value = (str(veg_dict[key[i]][0]) + veg_dict[key[i]][1]))
                    sheet.cell((x + i), 3).font = font0
                    sheet.cell((x + i), 2).font = font0
                    if len(veg_dict[key[i]]) > 2:
                        gg = 1
                        for j in range(len(veg_dict[key[i]])-2):
                            sheet.cell(row = (x + i), column = (4 + j), value = veg_dict[key[i]][2 + j])
                            sheet.cell((x + i), (4 + j)).font = font0
                    if key[i] in hit_dict:
                        for k in range(len(hit_dict[key[i]])):
                            if gg == 1:
                                sheet.cell(row = (x + i), column = (5 + j + k), value = hit_dict[key[i]][k])
                                sheet.cell((x + i), (5+j+k)).font = font0
                            else:
                                sheet.cell(row = (x + i), column = (4 + k), value = hit_dict[key[i]][k])
                                sheet.cell((x + i), (4 + k)).font = font0
                self.wb0.save('D:/菜單/0227菜單.xlsx')
            else:
                while True:
                    if sheet.cell(x, 2).value is None:
                        break
                    if sheet.cell(x, 2).value not in veg_dict:
                        if '斤' not in sheet.cell(x, 3).value:
                            veg_dict[sheet.cell(x, 2).value] = [0, sheet.cell(x, 3).value[-1]]
                        else:
                            veg_dict[sheet.cell(x, 2).value] = [0, 'fff']
                    try:
                        num = int(sheet.cell(x, 3).value[:len(sheet.cell(x, 3).value)-1])
                        veg_dict[sheet.cell(x, 2).value][0] += num
                    except:
                        veg_dict[sheet.cell(x, 2).value].append(sheet.cell(x, 3).value)
                    if sheet.cell(x, 4) is not None:
                        if sheet.cell(x, 2).value not in hit_dict:
                            hit_dict[sheet.cell(x, 2).value] = []
                        hit_dict[sheet.cell(x, 2).value].append(sheet.cell(x, 4).value)
                    x += 1


if __name__ == "__main__":
    app = Action()
    app.mainloop()
    